"""Excel output writer for processed fiscal documents."""

from __future__ import annotations

import re
import shutil
import unicodedata
import os
import posixpath
import zipfile
import xml.etree.ElementTree as ET
from copy import deepcopy
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Iterable, Literal

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lumina_bot.core.logger import get_logger
from lumina_bot.models.nota import NotaFiscal


ExcelOutputMode = Literal["single_sheet", "multi_sheet", "one_file_per_pdf"]


class ExcelWriter:
    """Write fiscal document rows to Excel workbooks."""

    TEMPLATE_FILENAME = "Lote_de_Fatura_CEF_Consignado.xlsx"
    TEMPLATE_SHEET = "Lançamentos"
    TEMPLATE_HEADER_ROW = 2
    TEMPLATE_DATA_ROW = 3
    TEMPLATE_DATA_COLUMNS = 185
    XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
    CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
    TEMPLATE_NAMESPACES = {
        "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
        "x15": "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main",
        "x15ac": "http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac",
        "x14": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
        "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
        "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
        "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
        "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
        "xr6": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision6",
        "xr10": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10",
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    }
    CALC_CHAIN_REL_TYPE = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain"
    )
    XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"
    ROOT_TAG_PATTERN = re.compile(
        rb"<(?:[A-Za-z_][\w.-]*:)?[A-Za-z_][\w.-]*(?:\s[^<>]*?)?>"
    )
    NAMESPACE_PATTERN = re.compile(
        rb"\s+xmlns(?::(?P<prefix>[A-Za-z_][\w.-]*))?=(?P<quote>['\"])(?P<uri>.*?)(?P=quote)"
    )

    SUMMARY_COLUMNS = [
        "Arquivo",
        "Tipo Documento",
        "Parser",
        "Sub Layout",
        "OCR Usado",
        "Número",
        "Série",
        "Modelo",
        "Chave",
        "Chave Acesso Raw",
        "Data Emissão",
        "Competência",
        "RPS Número",
        "Município Emissor NFS-e",
        "Município Incidência ISS",
        "Código NBS",
        "Código Obra",
        "Código CEI/CNO",
        "SFOBRAS",
        "Valor Aproximado Tributos Raw",
        "CNPJ Prestador",
        "CPF Prestador",
        "Razão Social Prestador",
        "Nome Fantasia Prestador",
        "CNAE",
        "Inscrição Municipal Prestador",
        "Inscrição Estadual Prestador",
        "Cidade Prestador",
        "UF Prestador",
        "CNPJ Tomador",
        "CPF Tomador",
        "Razão Social Tomador",
        "Nome Fantasia Tomador",
        "Cidade Tomador",
        "UF Tomador",
        "Código Serviço",
        "Descrição Serviço",
        "Discriminação",
        "Valor Bruto",
        "Valor Líquido",
        "Valor Total",
        "Base Cálculo",
        "Alíquota",
        "ISS",
        "INSS",
        "PIS",
        "COFINS",
        "CSLL",
        "IRRF",
        "Retenções",
        "Descontos",
        "Observações",
        "Status",
        "Erro",
        "SHA256",
        "Caminho Local",
        "Caminho Remoto",
    ]

    MONEY_KEYWORDS = (
        "valor",
        "iss",
        "inss",
        "pis",
        "cofins",
        "csll",
        "irrf",
        "retencoes",
        "descontos",
        "base_calculo",
        "aliquota",
    )
    DATE_KEYWORDS = ("data", "competencia")

    def __init__(self, output_path: Path) -> None:
        self._output_path = output_path
        self._logger = get_logger(self.__class__.__name__)

    def append(self, notas: Iterable[NotaFiscal]) -> None:
        """Append rows to a single-sheet workbook without dropping existing rows."""
        self.write(notas, mode="single_sheet")

    def write(
        self,
        notas: Iterable[NotaFiscal],
        *,
        mode: ExcelOutputMode = "single_sheet",
    ) -> None:
        """Write notes using the selected Excel output mode."""
        rows = list(notas)

        if not rows:
            return

        if mode == "multi_sheet":
            self.write_multi_sheet(rows)
            return

        if mode == "one_file_per_pdf":
            self.write_one_file_per_pdf(rows)
            return

        self.write_template(rows)

    def write_template(self, notas: Iterable[NotaFiscal]) -> None:
        """Fill the supplied accounting template without rebuilding its workbook.

        The template contains formulas, named cells, validations and support tabs
        required by the Lumina import.  Recreating it with a dataframe would lose
        those details, so only values in the data rows of ``Lançamentos`` are
        changed inside the original XLSX package.
        """
        rows = list(notas)

        if not rows:
            return

        template_path = self._template_path()
        if not template_path.is_file():
            raise FileNotFoundError(
                "Excel template not found. Set LINKAI_EXCEL_TEMPLATE_PATH or "
                f"install {self.TEMPLATE_FILENAME} in lumina_bot/templates."
            )

        self._output_path.parent.mkdir(parents=True, exist_ok=True)
        temporary_path = self._output_path.with_name(
            f".{self._output_path.name}.template.tmp"
        )

        try:
            self._fill_template(template_path, temporary_path, rows)
            os.replace(temporary_path, self._output_path)
        finally:
            temporary_path.unlink(missing_ok=True)

        self._logger.info(
            "Template Excel generated: %s | rows=%s | template=%s",
            self._output_path,
            len(rows),
            template_path,
        )

    def _template_path(self) -> Path:
        configured = os.getenv("LINKAI_EXCEL_TEMPLATE_PATH")
        if configured:
            return Path(configured).expanduser()

        return Path(__file__).resolve().parent.parent / "templates" / self.TEMPLATE_FILENAME

    def _fill_template(
        self,
        template_path: Path,
        output_path: Path,
        notas: list[NotaFiscal],
    ) -> None:
        namespace = {"m": self.XML_NS}
        for prefix, uri in self.TEMPLATE_NAMESPACES.items():
            ET.register_namespace(prefix, uri)
        ET.register_namespace("", self.XML_NS)
        ET.register_namespace("r", self.REL_NS)

        with zipfile.ZipFile(template_path, "r") as source:
            package = {name: source.read(name) for name in source.namelist()}

        workbook_root = ET.fromstring(package["xl/workbook.xml"])
        workbook_rels = ET.fromstring(package["xl/_rels/workbook.xml.rels"])
        sheet_element = next(
            sheet
            for sheet in workbook_root.findall("m:sheets/m:sheet", namespace)
            if sheet.attrib.get("name") == self.TEMPLATE_SHEET
        )
        relationship_id = sheet_element.attrib[f"{{{self.REL_NS}}}id"]
        relationship = next(
            rel
            for rel in workbook_rels
            if rel.attrib.get("Id") == relationship_id
        )
        target = relationship.attrib["Target"]
        sheet_path = (
            target.lstrip("/")
            if target.startswith("/")
            else posixpath.normpath(posixpath.join("xl", target))
        )

        sheet_root = ET.fromstring(package[sheet_path])
        sheet_data = sheet_root.find("m:sheetData", namespace)
        if sheet_data is None:
            raise ValueError(f"Template sheet {self.TEMPLATE_SHEET!r} has no sheetData.")

        template_rows = {
            int(row.attrib["r"]): row
            for row in sheet_data.findall("m:row", namespace)
            if row.attrib.get("r", "").isdigit()
        }
        if self.TEMPLATE_DATA_ROW not in template_rows:
            raise ValueError(
                f"Template sheet {self.TEMPLATE_SHEET!r} has no data row "
                f"{self.TEMPLATE_DATA_ROW}."
            )

        for row_number, row in template_rows.items():
            if row_number >= self.TEMPLATE_DATA_ROW:
                self._clear_row(row, namespace)

        # The accounting template receives one launch row per fiscal note.
        # Product details remain available in the normalized/XML output, but
        # must not duplicate the note total in the Lumina import sheet.
        values = [{} for _nota in notas]
        for value_index, nota in enumerate(notas):
            values[value_index] = self._template_row_values(
                nota,
                item=None,
                include_installments=True,
            )

        last_template_row = max(template_rows)
        for offset, row_values in enumerate(values):
            row_number = self.TEMPLATE_DATA_ROW + offset
            row = template_rows.get(row_number)
            if row is None:
                row = self._clone_template_row(
                    template_rows[last_template_row],
                    row_number,
                    namespace,
                )
                sheet_data.append(row)
            self._write_row_values(row, row_values, namespace)

        package[sheet_path] = self._serialize_template_xml(package[sheet_path], sheet_root)

        # The template contains a calculation chain for its sample formulas.
        # Data rows are replaced with values above, so keeping that stale chain
        # makes Excel report that the generated workbook needs repair.
        self._remove_calculation_chain(package, workbook_rels)

        calc_pr = workbook_root.find("m:calcPr", namespace)
        if calc_pr is None:
            calc_pr = ET.SubElement(workbook_root, f"{{{self.XML_NS}}}calcPr")
        calc_pr.set("calcMode", "auto")
        calc_pr.set("fullCalcOnLoad", "1")
        calc_pr.set("forceFullCalc", "1")
        package["xl/workbook.xml"] = self._serialize_template_xml(
            package["xl/workbook.xml"],
            workbook_root,
        )
        package["xl/_rels/workbook.xml.rels"] = ET.tostring(
            workbook_rels,
            encoding="utf-8",
            xml_declaration=True,
        )

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as destination:
            for name, content in package.items():
                destination.writestr(name, content)

    @classmethod
    def _remove_calculation_chain(
        cls,
        package: dict[str, bytes],
        workbook_rels: ET.Element,
    ) -> None:
        """Remove stale calculation-chain parts after replacing template formulas."""
        rel_namespace = {"r": cls.PACKAGE_REL_NS}
        calculation_relationships = [
            relationship
            for relationship in workbook_rels.findall("r:Relationship", rel_namespace)
            if relationship.attrib.get("Type") == cls.CALC_CHAIN_REL_TYPE
            or relationship.attrib.get("Target", "").endswith("calcChain.xml")
        ]
        for relationship in calculation_relationships:
            workbook_rels.remove(relationship)

        for name in list(package):
            if name == "xl/calcChain.xml" or name.endswith("/calcChain.xml"):
                package.pop(name, None)

        content_types = ET.fromstring(package["[Content_Types].xml"])
        content_namespace = {"ct": cls.CONTENT_TYPES_NS}
        for override in content_types.findall("ct:Override", content_namespace):
            if override.attrib.get("PartName") == "/xl/calcChain.xml":
                content_types.remove(override)

        package["[Content_Types].xml"] = ET.tostring(
            content_types,
            encoding="utf-8",
            xml_declaration=True,
        )

    @classmethod
    def _serialize_template_xml(
        cls,
        source_xml: bytes,
        root: ET.Element,
    ) -> bytes:
        """Serialize a template part without dropping its namespace declarations."""
        serialized = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        source_root = cls.ROOT_TAG_PATTERN.search(source_xml)
        output_root = cls.ROOT_TAG_PATTERN.search(serialized)

        if source_root is None or output_root is None:
            return serialized

        source_namespaces = {
            match.group("prefix") or b"": match.group("uri")
            for match in cls.NAMESPACE_PATTERN.finditer(source_root.group(0))
        }
        output_namespaces = {
            match.group("prefix") or b"": match.group("uri")
            for match in cls.NAMESPACE_PATTERN.finditer(output_root.group(0))
        }
        missing = b"".join(
            b' xmlns'
            + (b":" + prefix if prefix else b"")
            + b'="'
            + uri
            + b'"'
            for prefix, uri in source_namespaces.items()
            if prefix not in output_namespaces
        )

        if not missing:
            return serialized

        insertion_point = output_root.end() - 1
        return serialized[:insertion_point] + missing + serialized[insertion_point:]

    def _template_row_values(
        self,
        nota: NotaFiscal,
        *,
        item: Any | None = None,
        include_installments: bool = True,
    ) -> dict[str, Any]:
        """Map normalized fiscal data to the model's fixed launch columns."""
        extras = nota.outros_campos
        amount = (
            getattr(item, "valor_total", None)
            if item is not None
            else None
        )
        if amount is None:
            amount = nota.valor_total or nota.valor_bruto

        installments = nota.parcelas if include_installments else []
        first_installment = installments[0] if installments else None
        second_installment = installments[1] if len(installments) > 1 else None
        service_code = nota.codigo_servico

        values: dict[str, Any] = {
            "A": self._configured_or_extra(
                extras,
                "LINKAI_TEMPLATE_BILLING_CNPJ",
                "cnpj_faturamento",
                "cost_collector_client_cnpj",
            ),
            "B": self._configured_or_extra(
                extras,
                "LINKAI_TEMPLATE_BILLING_CLIENT",
                "cliente_faturamento",
                "cost_collector_client_trade_name",
            ),
            "C": self._extra_from_mapping(extras, "socio", "partner_name"),
            "D": self._extra_from_mapping(extras, "cc", "codigo_cc", "centro_custo"),
            "E": nota.prestador.cnpj or nota.prestador.cpf,
            "F": nota.prestador.razao_social or nota.prestador.nome_fantasia,
            "G": self._extra_from_mapping(
                extras,
                "codigo_fornecedor_lumina",
                "cod_fornecedor_lumina",
                "supplier_id",
            ),
            "H": nota.numero,
            "I": nota.serie,
            "J": self._date_value(nota.data_emissao),
            "K": service_code,
            # These columns require internal Lumina data that is not present
            # in fiscal PDFs. They must remain blank until supplied by Lumina.
            "L": None,
            "M": None,
            "N": amount,
            "O": self._date_value(first_installment.vencimento) if first_installment else None,
            "Q": first_installment.valor if first_installment else None,
            "R": None,
            "S": first_installment.valor if first_installment else None,
            "AC": self._date_value(second_installment.vencimento) if second_installment else None,
            "AD": second_installment.valor if second_installment else None,
            "AE": None,
            "AF": second_installment.valor if second_installment else None,
            "AP": 0,
            "AQ": self._net_value(nota, amount),
            "CE": nota.observacoes,
        }

        iss = nota.tributos.iss
        is_nfe = nota.modelo == "55" or nota.tipo_documento == "NFE_DANFE_55"
        if not is_nfe and (iss is not None or nota.tributos.base_calculo is not None):
            values.update(
                {
                    "AR": self._extra_from_mapping(extras, "iss_tipo", "iss_retido"),
                    "AS": self._extra_from_mapping(extras, "iss_codigo", "codigo_servico") or service_code,
                    "AT": nota.tributos.base_calculo,
                    "AU": nota.tributos.aliquota,
                    "AV": iss,
                    "AW": self._extra_from_mapping(extras, "iss_codigo_prefeitura", "codigo_prefeitura"),
                    "AX": self._extra_from_mapping(extras, "iss_prefeitura", "prefeitura"),
                    "AY": nota.municipio or nota.prestador.endereco.cidade,
                    "AZ": self._date_value(self._extra_from_mapping(extras, "iss_data_pagamento", "data_pagamento_iss")),
                }
            )

        for prefix, code_col, base_col, rate_col, value_col, date_col in (
            ("inss", "BA", "BB", "BC", "BD", "BE"),
            ("irrf", "BF", "BG", "BH", "BI", "BJ"),
            ("pcc", "BK", "BL", "BM", "BN", "BO"),
            ("pis", "BP", "BQ", "BR", "BS", "BT"),
            ("cofins", "BU", "BV", "BW", "BX", "BY"),
            ("csll", "BZ", "CA", "CB", "CC", "CD"),
        ):
            tax_value = self._tax_value(nota, prefix)
            tax_base = self._extra_from_mapping(extras, f"{prefix}_base", f"{prefix}_base_value")
            tax_rate = self._extra_from_mapping(extras, f"{prefix}_aliquota", f"{prefix}_rate")
            tax_code = self._extra_from_mapping(extras, f"{prefix}_codigo", f"{prefix}_code")
            tax_date = self._extra_from_mapping(extras, f"{prefix}_data_pagamento", f"{prefix}_due_date")
            if any(value is not None for value in (tax_value, tax_base, tax_rate, tax_code, tax_date)):
                values.update(
                    {
                        code_col: tax_code,
                        base_col: tax_base,
                        rate_col: tax_rate,
                        value_col: tax_value,
                        date_col: self._date_value(tax_date),
                    }
                )

        allocations = extras.get("alocacoes") or extras.get("pro_ratings") or []
        if isinstance(allocations, list):
            for index, allocation in enumerate(allocations[:50]):
                if not isinstance(allocation, dict):
                    continue
                id_column = self._column_after("CG", index * 2)
                value_column = self._column_after("CG", index * 2 + 1)
                values[id_column] = allocation.get("id") or allocation.get("project_task_id")
                values[value_column] = allocation.get("value") or allocation.get("valor")

        return {column: value for column, value in values.items() if value is not None}

    def _tax_value(self, nota: NotaFiscal, prefix: str) -> Any:
        if prefix == "pcc":
            return self._extra_from_mapping(
                nota.outros_campos,
                "pcc",
                "pcc_value",
                "pis_cofins",
            )
        return getattr(nota.tributos, prefix, None)

    def _net_value(self, nota: NotaFiscal, amount: Any) -> Any:
        if amount is None:
            return None
        deductions = [nota.tributos.iss]
        deductions.extend(
            self._tax_value(nota, prefix)
            for prefix in ("inss", "irrf", "pcc", "pis", "cofins", "csll")
        )
        return float(amount) - sum(float(value) for value in deductions if isinstance(value, (int, float)))

    @staticmethod
    def _extra_from_mapping(mapping: dict[str, Any], *keys: str) -> Any:
        for key in keys:
            value = mapping.get(key)
            if value is not None and value != "":
                return value
        return None

    @classmethod
    def _configured_or_extra(cls, mapping: dict[str, Any], *keys: str) -> Any:
        for key in keys:
            if key.startswith("LINKAI_"):
                value = os.getenv(key)
                if value:
                    return value
            else:
                value = mapping.get(key)
                if value is not None and value != "":
                    return value
        return None

    @staticmethod
    def _date_value(value: Any) -> date | datetime | None:
        if value is None or value == "":
            return None
        if isinstance(value, (date, datetime)):
            return value
        text = str(value).strip()
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            match = re.search(r"(\d{2})/(\d{2})/(\d{4})", text)
            if match:
                day, month, year = match.groups()
                return date(int(year), int(month), int(day))
        return None

    @staticmethod
    def _column_after(start: str, offset: int) -> str:
        number = 0
        for character in start:
            number = number * 26 + ord(character.upper()) - 64
        number += offset
        result = ""
        while number:
            number, remainder = divmod(number - 1, 26)
            result = chr(65 + remainder) + result
        return result

    @classmethod
    def _clear_row(cls, row: ET.Element, namespace: dict[str, str]) -> None:
        for cell in row.findall("m:c", namespace):
            cell.attrib.pop("t", None)
            for child in list(cell):
                if child.tag.rsplit("}", 1)[-1] in {"f", "v", "is"}:
                    cell.remove(child)

    @classmethod
    def _clone_template_row(
        cls,
        source_row: ET.Element,
        row_number: int,
        namespace: dict[str, str],
    ) -> ET.Element:
        row = deepcopy(source_row)
        row.attrib["r"] = str(row_number)
        for cell in row.findall("m:c", namespace):
            reference = cell.attrib.get("r", "")
            column = re.sub(r"\d+$", "", reference)
            cell.attrib["r"] = f"{column}{row_number}"
        cls._clear_row(row, namespace)
        return row

    @classmethod
    def _write_row_values(
        cls,
        row: ET.Element,
        values: dict[str, Any],
        namespace: dict[str, str],
    ) -> None:
        cells = {
            re.sub(r"\d+$", "", cell.attrib.get("r", "")): cell
            for cell in row.findall("m:c", namespace)
        }
        for column, value in values.items():
            cell = cells.get(column)
            if cell is None:
                cell = ET.SubElement(row, f"{{{cls.XML_NS}}}c", {"r": f"{column}{row.attrib['r']}"})
                cells[column] = cell
            cls._set_cell_value(cell, value)

    @classmethod
    def _set_cell_value(cls, cell: ET.Element, value: Any) -> None:
        cell.attrib.pop("t", None)
        for child in list(cell):
            if child.tag.rsplit("}", 1)[-1] in {"f", "v", "is"}:
                cell.remove(child)

        if value is None:
            return

        if isinstance(value, (datetime, date)):
            if isinstance(value, datetime):
                timestamp = value.replace(tzinfo=None)
            else:
                timestamp = datetime.combine(value, datetime.min.time())
            value = (timestamp - datetime(1899, 12, 30)).total_seconds() / 86400

        if isinstance(value, bool):
            cell.attrib["t"] = "b"
            serialized = "1" if value else "0"
            value_node = ET.SubElement(cell, f"{{{cls.XML_NS}}}v")
            value_node.text = serialized
            return

        if isinstance(value, (int, float)) and not isinstance(value, bool):
            value_node = ET.SubElement(cell, f"{{{cls.XML_NS}}}v")
            value_node.text = f"{value:.15g}"
            return

        cell.attrib["t"] = "inlineStr"
        inline = ET.SubElement(cell, f"{{{cls.XML_NS}}}is")
        text_node = ET.SubElement(inline, f"{{{cls.XML_NS}}}t")
        text_node.text = str(value)
        if text_node.text[:1].isspace() or text_node.text[-1:].isspace():
            text_node.set(cls.XML_SPACE, "preserve")

    def write_structured(self, notas: Iterable[NotaFiscal]) -> None:
        """Write the canonical workbook tabs without flattening child records."""
        rows = list(notas)
        if not rows:
            return

        documents = [self._summary_row(nota) | {
            "Layout": nota.layout,
            "Parcelas": len(nota.parcelas),
            "Validações": len(nota.validacoes),
        } for nota in rows]
        items: list[dict[str, Any]] = []
        installments: list[dict[str, Any]] = []
        taxes: list[dict[str, Any]] = []
        validations: list[dict[str, Any]] = []

        for nota in rows:
            identity = {
                "Arquivo": nota.arquivo,
                "Tipo Documento": nota.tipo_documento,
                "Layout": nota.layout,
                "Número": nota.numero,
                "Chave": nota.chave,
            }
            for item_index, item in enumerate(nota.itens, start=1):
                items.append(identity | {"Item": item_index} | item.to_dict())
            for parcela in nota.parcelas:
                installments.append(identity | parcela.to_dict())
            for name, value in nota.tributos.to_dict().items():
                if isinstance(value, dict):
                    for child_name, child_value in value.items():
                        taxes.append(identity | {"Tributo": f"{name}.{child_name}", "Valor": child_value})
                else:
                    taxes.append(identity | {"Tributo": name, "Valor": value})
            for validation in nota.validacoes:
                validations.append(identity | validation.to_dict())

        frames = {
            "documentos": pd.DataFrame(documents),
            "itens": pd.DataFrame(items, columns=[
                "Arquivo", "Tipo Documento", "Layout", "Número", "Chave", "Item",
                "codigo", "descricao", "ncm", "cfop", "unidade", "quantidade",
                "valor_unitario", "valor_desconto", "valor_total", "cst",
                "base_calculo_icms", "valor_icms", "valor_ipi", "aliquota_icms",
                "aliquota_ipi", "valor_total_tributos", "outros_campos",
            ]),
            "parcelas": pd.DataFrame(installments, columns=[
                "Arquivo", "Tipo Documento", "Layout", "Número", "Chave",
                "numero", "vencimento", "valor", "raw", "pagina",
            ]),
            "tributos": pd.DataFrame(taxes, columns=[
                "Arquivo", "Tipo Documento", "Layout", "Número", "Chave", "Tributo", "Valor",
            ]),
            "validacoes": pd.DataFrame(validations, columns=[
                "Arquivo", "Tipo Documento", "Layout", "Número", "Chave", "regra", "status",
                "valor_extraido", "valor_calculado", "mensagem",
            ]),
        }
        self._write_frames(frames)

    def _write_frames(self, frames: dict[str, pd.DataFrame], output_path: Path | None = None) -> None:
        path = output_path or self._output_path
        path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_name, frame in frames.items():
                frame.to_excel(writer, index=False, sheet_name=sheet_name)
        self._format_workbook(path)
        self._logger.info("Structured Excel updated: %s | sheets=%s", path, len(frames))

    def write_single_sheet(self, notas: Iterable[NotaFiscal]) -> None:
        """Write all documents using the fixed accounting template."""
        self.write_template(notas)

    def write_multi_sheet(
        self,
        notas: Iterable[NotaFiscal],
        sheet_name_resolver: Callable[[NotaFiscal], str] | None = None,
    ) -> None:
        """Write all documents into one workbook with one sheet per document."""
        rows = list(notas)

        if not rows:
            return

        self._output_path.parent.mkdir(parents=True, exist_ok=True)
        resolver = sheet_name_resolver or self._default_sheet_name
        used_names: set[str] = set()

        with pd.ExcelWriter(self._output_path, engine="openpyxl") as writer:
            for nota in rows:
                sheet_name = self._unique_sheet_name(resolver(nota), used_names)
                pd.DataFrame(
                    [self._summary_row(nota)],
                    columns=self.SUMMARY_COLUMNS,
                ).to_excel(
                    writer,
                    index=False,
                    sheet_name=sheet_name,
                )

        self._format_workbook()
        self._logger.info(
            "Multi-sheet Excel updated: %s | sheets=%s",
            self._output_path,
            len(rows),
        )

    def write_one_file_per_pdf(self, notas: Iterable[NotaFiscal]) -> None:
        """Create one Excel file for each processed document."""
        rows = list(notas)

        if not rows:
            return

        output_dir = self._output_path.parent
        output_dir.mkdir(parents=True, exist_ok=True)

        for nota in rows:
            file_name = self._safe_file_name(self._default_sheet_name(nota)) + ".xlsx"
            file_path = self._unique_file_path(output_dir / file_name)

            temporary_writer = ExcelWriter(file_path)
            temporary_writer.write_template([nota])

        self._logger.info("One-file-per-PDF Excel export completed: %s", output_dir)

    def _format_workbook(self, path: Path | None = None) -> None:
        workbook_path = path or self._output_path

        if not workbook_path.is_file():
            return

        workbook = load_workbook(workbook_path)

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="F92B70",
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            sheet.row_dimensions[1].height = 24

            for column_cells in sheet.columns:
                first_cell = column_cells[0]
                column_name = str(first_cell.value or "")
                column_letter = get_column_letter(first_cell.column)
                max_length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )
                sheet.column_dimensions[column_letter].width = min(
                    max(max_length + 2, 12),
                    80,
                )

                for cell in column_cells[1:]:
                    cell.alignment = Alignment(vertical="top", wrap_text=False)

                if self._is_money_column(column_name):
                    for cell in column_cells[1:]:
                        cell.number_format = '#,##0.00'

                if self._is_date_column(column_name):
                    for cell in column_cells[1:]:
                        cell.number_format = "dd/mm/yyyy"

        workbook.save(workbook_path)

    def _default_sheet_name(self, nota: NotaFiscal) -> str:
        candidates = (
            nota.prestador.nome_fantasia,
            nota.prestador.razao_social,
            nota.tomador.razao_social,
            nota.numero,
            nota.arquivo,
            "Nota",
        )

        for candidate in candidates:
            if candidate:
                return str(candidate)

        return "Nota"

    def _summary_row(self, nota: NotaFiscal) -> dict[str, Any]:
        return {
            "Arquivo": nota.arquivo,
            "Tipo Documento": nota.tipo_documento,
            "Parser": nota.parser,
            "Sub Layout": nota.sub_layout,
            "OCR Usado": nota.ocr_used,
            "Número": nota.numero,
            "Série": nota.serie,
            "Modelo": nota.modelo,
            "Chave": nota.chave,
            "Chave Acesso Raw": nota.chave_acesso_raw,
            "Data Emissão": nota.data_emissao,
            "Competência": nota.competencia,
            "RPS Número": nota.rps_numero,
            "Município Emissor NFS-e": nota.municipio_emissor_nfse,
            "Município Incidência ISS": nota.municipio_incidencia_iss,
            "Código NBS": nota.codigo_nbs,
            "Código Obra": nota.codigo_obra,
            "Código CEI/CNO": nota.codigo_cei_cno,
            "SFOBRAS": nota.sfo_bras,
            "Valor Aproximado Tributos Raw": nota.valor_aproximado_tributos_raw,
            "CNPJ Prestador": nota.prestador.cnpj,
            "CPF Prestador": nota.prestador.cpf,
            "Razão Social Prestador": nota.prestador.razao_social,
            "Nome Fantasia Prestador": nota.prestador.nome_fantasia,
            "CNAE": self._extra_value(nota, "cnae", "codigo_cnae", "cnae_fiscal"),
            "Inscrição Municipal Prestador": nota.prestador.inscricao_municipal,
            "Inscrição Estadual Prestador": nota.prestador.inscricao_estadual,
            "Cidade Prestador": nota.prestador.endereco.cidade,
            "UF Prestador": nota.prestador.endereco.uf,
            "CNPJ Tomador": nota.tomador.cnpj,
            "CPF Tomador": nota.tomador.cpf,
            "Razão Social Tomador": nota.tomador.razao_social,
            "Nome Fantasia Tomador": nota.tomador.nome_fantasia,
            "Cidade Tomador": nota.tomador.endereco.cidade,
            "UF Tomador": nota.tomador.endereco.uf,
            "Código Serviço": nota.codigo_servico,
            "Descrição Serviço": nota.descricao_servico,
            "Discriminação": nota.discriminacao,
            "Valor Bruto": nota.valor_bruto,
            "Valor Líquido": nota.valor_liquido,
            "Valor Total": nota.valor_total,
            "Base Cálculo": nota.tributos.base_calculo,
            "Alíquota": nota.tributos.aliquota,
            "ISS": nota.tributos.iss,
            "INSS": nota.tributos.inss,
            "PIS": nota.tributos.pis,
            "COFINS": nota.tributos.cofins,
            "CSLL": nota.tributos.csll,
            "IRRF": nota.tributos.irrf,
            "Retenções": nota.tributos.retencoes,
            "Descontos": nota.tributos.descontos,
            "Observações": nota.observacoes,
            "Status": nota.status_processamento,
            "Erro": nota.erro_processamento,
            "SHA256": nota.sha256,
            "Caminho Local": nota.caminho_local,
            "Caminho Remoto": nota.caminho_remoto,
        }

    def _backup_incompatible_workbook(self) -> None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self._output_path.with_name(
            f"{self._output_path.stem}_backup_layout_antigo_{timestamp}{self._output_path.suffix}"
        )
        shutil.copy2(self._output_path, backup_path)
        self._logger.info(
            "Existing Excel used an old layout and was backed up: %s",
            backup_path,
        )

    @classmethod
    def _is_compatible_summary_frame(cls, frame: pd.DataFrame) -> bool:
        return list(frame.columns) == cls.SUMMARY_COLUMNS

    def _unique_file_path(self, desired_path: Path) -> Path:
        if not desired_path.exists():
            return desired_path

        counter = 2
        candidate = desired_path.with_name(
            f"{desired_path.stem}_{counter}{desired_path.suffix}"
        )

        while candidate.exists():
            counter += 1
            candidate = desired_path.with_name(
                f"{desired_path.stem}_{counter}{desired_path.suffix}"
            )

        return candidate

    def _extra_value(self, nota: NotaFiscal, *keys: str) -> Any:
        for key in keys:
            direct = nota.outros_campos.get(key)

            if direct:
                return direct

        xml_values = nota.outros_campos.get("xml_campos_extraidos")

        if isinstance(xml_values, dict):
            for key in keys:
                value = xml_values.get(key)

                if value:
                    return value

        return None

    def _unique_sheet_name(self, desired_name: str, used_names: set[str]) -> str:
        base = self._safe_sheet_name(desired_name)
        name = base
        suffix = 2

        while name in used_names:
            suffix_text = f"_{suffix}"
            name = f"{base[:31 - len(suffix_text)]}{suffix_text}"
            suffix += 1

        used_names.add(name)
        return name

    @staticmethod
    def _safe_sheet_name(value: str) -> str:
        cleaned = re.sub(r"[\[\]:*?/\\]", " ", value).strip()
        cleaned = re.sub(r"\s+", " ", cleaned)
        return (cleaned or "Nota")[:31]

    @staticmethod
    def _safe_file_name(value: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9_. -]+", "_", value).strip(" ._")
        return cleaned or "nota"

    def _is_money_column(self, column_name: str) -> bool:
        normalized = self._normalize_column_name(column_name)
        return any(keyword in normalized for keyword in self.MONEY_KEYWORDS)

    def _is_date_column(self, column_name: str) -> bool:
        normalized = self._normalize_column_name(column_name)
        return any(keyword in normalized for keyword in self.DATE_KEYWORDS)

    @staticmethod
    def _normalize_column_name(column_name: str) -> str:
        normalized = unicodedata.normalize("NFKD", column_name)
        normalized = "".join(
            character for character in normalized if not unicodedata.combining(character)
        )
        return normalized.lower().replace(" ", "_")
