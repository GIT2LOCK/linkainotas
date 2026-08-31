"""Tests for the fixed Lumina accounting workbook export."""

from __future__ import annotations

import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from lumina_bot.core.excel_writer import ExcelWriter
from lumina_bot.models.emitente import Emitente
from lumina_bot.models.item import Item
from lumina_bot.models.nota import NotaFiscal
from lumina_bot.models.parcela import Parcela


class ExcelTemplateTests(unittest.TestCase):
    def test_export_preserves_template_and_clears_sample_rows(self) -> None:
        note = NotaFiscal(
            arquivo="DANFE 123.pdf",
            numero="0000123",
            serie="1",
            data_emissao="2026-08-20",
            prestador=Emitente(
                cnpj="12.345.678/0001-90",
                razao_social="Fornecedor de Teste LTDA",
            ),
            valor_bruto=150.0,
            valor_total=150.0,
            itens=[
                Item(
                    codigo="MAT-01",
                    descricao="Material de teste",
                    valor_total=150.0,
                )
            ],
            parcelas=[
                Parcela(numero="001", vencimento="2026-09-20", valor=150.0)
            ],
        )

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "notas.xlsx"
            ExcelWriter(output).write([note])

            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                ["Lançamentos", "Planilha1", "Encargos", "Impostos", "NomeCelulas"],
            )
            launches = workbook["Lançamentos"]
            self.assertEqual(launches["H2"].value, "NR NF")
            self.assertEqual(launches["E3"].value, "12.345.678/0001-90")
            self.assertEqual(launches["F3"].value, "Fornecedor de Teste LTDA")
            self.assertEqual(launches["H3"].value, "0000123")
            self.assertEqual(launches["L3"].value, "MAT-01")
            self.assertEqual(launches["N3"].value, 150)
            self.assertEqual(launches["Q3"].value, 150)
            self.assertEqual(launches["O3"].value, datetime(2026, 9, 20))
            self.assertIsNone(launches["E4"].value)
            self.assertIsNone(launches["F4"].value)
            self.assertIsNone(launches["H4"].value)
            self.assertEqual(workbook["Planilha1"]["B1"].value, "=A1/A13")

            with ZipFile(output) as archive:
                self.assertNotIn("xl/calcChain.xml", archive.namelist())

                relationships = ET.fromstring(
                    archive.read("xl/_rels/workbook.xml.rels")
                )
                self.assertFalse(
                    any(
                        relationship.attrib.get("Target", "").endswith("calcChain.xml")
                        for relationship in relationships
                    )
                )

                content_types = ET.fromstring(archive.read("[Content_Types].xml"))
                self.assertFalse(
                    any(
                        override.attrib.get("PartName") == "/xl/calcChain.xml"
                        for override in content_types
                    )
                )


if __name__ == "__main__":
    unittest.main()
